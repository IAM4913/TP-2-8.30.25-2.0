from openpyxl import load_workbook
import json, datetime, pathlib
root = r"c:\Users\micha\Documents\projects\tod\Truck Planner\Truck Planner 2 8.30.25"
wb_path = pathlib.Path(root)/"_dh_load_list.xlsx"
opt_path = pathlib.Path(root)/"optimize.json"
print('READING', wb_path)
wb = load_workbook(str(wb_path))
for sheet in wb.sheetnames:
    ws = wb[sheet]
    headers = [cell.value for cell in ws[1]]
    print('\nSHEET:', sheet)
    print('HEADERS:', headers)
    print('FIRST 5 ROWS:')
    row_count = 0
    for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
        print(row[:12])
        row_count += 1
        if row_count >= 5:
            break

print('\nREADING', opt_path)
with open(opt_path, 'r', encoding='utf-8') as f:
    data = json.load(f)

results = []

def walk(x):
    if isinstance(x, dict):
        pk = x.get('priorityBucket') or x.get('priority_bucket')
        ed = x.get('earliestDue') or x.get('earliest_due') or x.get('earliest')
        if pk == 'Late' and ed:
            parsed = None
            try:
                parsed = datetime.datetime.fromisoformat(ed)
            except Exception:
                try:
                    parsed = datetime.datetime.strptime(ed[:10], '%Y-%m-%d')
                except Exception:
                    parsed = None
            if parsed is not None:
                tomorrow = datetime.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) + datetime.timedelta(days=1)
                if parsed > tomorrow:
                    results.append({'so': x.get('so'), 'line': x.get('line'), 'earliestDue': ed})
        for v in x.values():
            walk(v)
    elif isinstance(x, list):
        for i in x:
            walk(i)

walk(data)
print('\nLate assignments with earliestDue > tomorrow:', len(results))
for r in results[:50]:
    print(r)
