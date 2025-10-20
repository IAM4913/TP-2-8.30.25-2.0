from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any, cast
import datetime as _dt
import pandas as pd
from io import BytesIO
import re
import math
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from .schemas import (
    UploadPreviewResponse,
    OptimizeRequest,
    OptimizeResponse,
    TruckSummary,
    LineAssignment,
    CombineTrucksRequest,
    CombineTrucksResponse,
    DemandSite,
    DepotLeg,
    SingleStopTruck,
    OptimizeV1Response,
)
from .excel_utils import compute_calculated_fields, _find_planning_whse_col, filter_by_planning_whse, extract_unique_addresses, apply_routing_filters
from .optimizer_simple import naive_grouping, NO_MULTI_STOP_CUSTOMERS
from dotenv import load_dotenv  # new
import os
import psycopg
from urllib.parse import quote_plus, urlparse, urlunparse, parse_qsl, urlencode
from .geocode_service import (
    google_geocode_query,
    cache_lookup_address,
    cache_upsert_address,
    build_address_query,
)
from .distance_service import google_distance_matrix, haversine_matrix
from .vrp_solver import Stop, Route, solve_vrp


REQUIRED_COLUMNS_MAPPED: List[str] = [
    "SO",  # Sales Order Number
    "Line",  # Line Number
    "Customer",  # Customer Name
    "shipping_city",  # Shipping City
    "shipping_state",  # Shipping State
    "Ready Weight",  # Ready Weight (lbs)
    "RPcs",  # Ready Pieces (quantity)
    "Grd",  # Material Grade
    "Size",  # Material Thickness (Size)
    "Width",  # Material Width
    "Earliest Due",  # Earliest Due Date
    "Latest Due",  # Latest Due Date
]


app = FastAPI(title="Truck Planner Backend", version="0.1.0")

# Enable CORS for local development UIs
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

load_dotenv()  # new: loads backend/.env


def _ensure_sslmode(dsn: str) -> str:
    """Append sslmode=require if not present in DSN query string.

    Handles both DSN with and without existing query params.
    """
    try:
        parsed = urlparse(dsn)
        # If not a URL (e.g., key=value DSN), just return as-is
        if not parsed.scheme or not parsed.netloc:
            return dsn
        q = dict(parse_qsl(parsed.query))
        if "sslmode" not in q:
            q["sslmode"] = "require"
            new_query = urlencode(q)
            return urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment))
        return dsn
    except Exception:
        return dsn


def _build_supabase_dsn() -> str:
    """Build a Postgres DSN.

    Priority:
    1) SUPABASE_DB_URL (if provided)
    2) Construct from SUPABASE_DB_HOST, SUPABASE_DB_PORT, SUPABASE_DB_NAME, SUPABASE_DB_USER, SUPABASE_DB_PASSWORD
       (password is URL-encoded to handle special characters)
    Always ensures sslmode=require for Supabase.
    """
    # 1) Direct URL
    raw = os.getenv("SUPABASE_DB_URL")
    if raw:
        return _ensure_sslmode(raw)

    # 2) Assemble from parts
    host = os.getenv("SUPABASE_DB_HOST")
    user = os.getenv("SUPABASE_DB_USER", "postgres")
    password = os.getenv("SUPABASE_DB_PASSWORD")
    dbname = os.getenv("SUPABASE_DB_NAME", "postgres")
    port = os.getenv("SUPABASE_DB_PORT", "5432")
    if host and password:
        pw_enc = quote_plus(password)
        dsn = f"postgresql://{user}:{pw_enc}@{host}:{port}/{dbname}?sslmode=require"
        return dsn
    return ""


def _init_geo_tables() -> None:
    """Create minimal tables for Phase 1 routing foundation if they don't exist.

    Tables:
      - address_cache: normalized address, raw components, lat/lng, confidence, provider, updated_at
      - customer_locations: customer key/name mapped to address_cache entry
      - distance_cache: cached distance/time between two keys (e.g., normalized addresses)
      - depot_config: configured depot location
    """
    dsn = _build_supabase_dsn()
    if not dsn:
        # Running without DB configured – skip silently per PRD graceful degradation
        print("[routing:init] No DB configured; skipping table initialization")
        return
    ddl_statements = [
        # Address cache keyed by normalized address
        (
            "address_cache",
            """
            create table if not exists address_cache (
                id bigserial primary key,
                normalized text unique not null,
                street text,
                suite text,
                city text,
                state text,
                zip text,
                country text default 'USA',
                latitude double precision,
                longitude double precision,
                confidence double precision,
                provider text,
                updated_at timestamp without time zone default now()
            );
            """,
        ),
        # Customer to address mapping
        (
            "customer_locations",
            """
            create table if not exists customer_locations (
                id bigserial primary key,
                customer_key text not null,
                normalized_address text not null,
                address_id bigint references address_cache(id) on delete set null,
                updated_at timestamp without time zone default now(),
                unique(customer_key)
            );
            """,
        ),
        # Distance cache between two normalized addresses (directed pair + provider)
        (
            "distance_cache",
            """
            create table if not exists distance_cache (
                origin_normalized text not null,
                dest_normalized text not null,
                provider text not null,
                distance_miles double precision,
                duration_minutes double precision,
                updated_at timestamp without time zone default now(),
                primary key (origin_normalized, dest_normalized, provider)
            );
            """,
        ),
        # Depot configuration (single row expected)
        (
            "depot_config",
            """
            create table if not exists depot_config (
                id smallint primary key default 1,
                name text,
                address text,
                latitude double precision,
                longitude double precision,
                updated_at timestamp without time zone default now()
            );
            """,
        ),
    ]
    try:
        with psycopg.connect(dsn, connect_timeout=5) as conn:
            with conn.cursor() as cur:
                for _name, ddl in ddl_statements:
                    cur.execute(ddl)
            conn.commit()
        print("[routing:init] Geo tables ensured")
    except Exception as exc:
        # Do not crash app; PRD requires graceful degradation without DB
        print(f"[routing:init] Failed to ensure tables: {exc}")


@app.get("/health")
def health() -> dict:
    return {"status": "ok"}


@app.get("/db/ping")
def db_ping() -> Dict[str, Any]:
    """Test connection to Supabase Postgres using SUPABASE_DB_URL.

    Returns a tiny payload with server_version and current_timestamp from the DB.
    """
    dsn = _build_supabase_dsn()
    if not dsn:
        raise HTTPException(
            status_code=500, detail="Database connection not configured. Set SUPABASE_DB_URL or SUPABASE_DB_HOST + SUPABASE_DB_PASSWORD.")
    try:
        with psycopg.connect(dsn, connect_timeout=5) as conn:
            with conn.cursor() as cur:
                cur.execute("select version(), now();")
                version, now_val = cur.fetchone()
        return {"ok": True, "version": str(version), "now": str(now_val)}
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"DB ping failed: {exc}") from exc


@app.post("/upload/preview", response_model=UploadPreviewResponse)
async def upload_preview(file: UploadFile = File(...)) -> UploadPreviewResponse:
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="Only .xlsx files are supported")

    try:
        content: bytes = await file.read()
        buffer = BytesIO(content)
        # Read with openpyxl engine for .xlsx
        df: pd.DataFrame = pd.read_excel(buffer, engine="openpyxl")
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=400, detail=f"Failed to read Excel: {exc}") from exc

    # Normalize headers by exact names as in the source file (no strip/rename yet)
    headers: List[str] = list(map(str, df.columns.tolist()))

    missing = [col for col in REQUIRED_COLUMNS_MAPPED if col not in headers]

    # Provide a small sample for UI validation
    sample_records = df.head(5).to_dict(
        orient="records") if not df.empty else []

    # Normalize sample keys to strings
    norm_sample: List[Dict[str, Any]] = []
    for rec in sample_records:
        norm_sample.append({str(k): v for k, v in rec.items()})

    return UploadPreviewResponse(
        headers=headers,
        rowCount=int(df.shape[0]),
        missingRequiredColumns=missing,
        sample=norm_sample,
    )


@app.post("/optimize", response_model=OptimizeResponse)
async def optimize(
    file: UploadFile = File(...),
    planningWhse: Optional[str] = Form("ZAC"),
) -> OptimizeResponse:
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="Only .xlsx files are supported")
    try:
        content: bytes = await file.read()
        buffer = BytesIO(content)
        df: pd.DataFrame = pd.read_excel(buffer, engine="openpyxl")
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=400, detail=f"Failed to read Excel: {exc}") from exc

    # Filter to Planning Whse first (defaults to ZAC); if column not present, no-op
    if planningWhse:
        try:
            df = filter_by_planning_whse(df, allowed_values=(planningWhse,))
        except Exception:
            # If anything goes wrong, fall back to unfiltered to avoid blocking
            pass

    df = compute_calculated_fields(df)
    df = apply_routing_filters(df)

    # Debug file write to confirm optimization started
    from pathlib import Path
    debug_file = Path(__file__).parent.parent / "LAST_RUN.txt"
    with open(debug_file, "w") as f:
        f.write(f"[/optimize endpoint] Started at {_dt.datetime.now()}\n")
        f.write(f"Order lines after filters: {len(df)}\n")
        f.write(f"Customers: {df['Customer'].nunique()}\n")
        f.write(f"Total weight: {df['Ready Weight'].sum():,.0f} lbs\n")
        f.write(
            f"Customer groups: {df.groupby(['Customer', 'shipping_city', 'shipping_state']).ngroups}\n")

    # Use default weight config for now - we'll add form parameter support later
    cfg = {
        "texas_max_lbs": 52000,
        "texas_min_lbs": 47000,
        "other_max_lbs": 48000,
        "other_min_lbs": 44000,
    }

    try:
        print(f"DataFrame shape: {df.shape}")
        print(f"Available columns: {list(df.columns)}")
        print(f"Weight config: {cfg}")
        print(
            f"[/optimize] Starting optimization with {len(df)} lines, {df['Customer'].nunique()} customers")
        print(f"[/optimize] Total weight: {df['Ready Weight'].sum():,.0f} lbs")
        print(
            f"[/optimize] Expected trucks at 48k lbs avg: ~{df['Ready Weight'].sum() / 48000:.1f}")
        trucks_df, assigns_df = naive_grouping(df, cfg)
        print(
            f"Optimization successful: {len(trucks_df)} trucks, {len(assigns_df)} assignments")

        # Update debug file with results
        with open(debug_file, "a") as f:
            f.write(f"\nOptimization completed:\n")
            f.write(f"Trucks created: {len(trucks_df)}\n")
            f.write(f"Assignments: {len(assigns_df)}\n")
            if not trucks_df.empty:
                f.write(
                    f"Total truck weight: {trucks_df['totalWeight'].sum():,.0f} lbs\n")
                f.write(
                    f"Avg truck weight: {trucks_df['totalWeight'].mean():,.0f} lbs\n")
    except Exception as exc:  # noqa: BLE001
        print(f"Optimization error: {exc}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=400, detail=f"Optimization failed: {exc}") from exc

    # Build sections mapping
    sections: Dict[str, List[int]] = {}
    if not trucks_df.empty:
        for bucket, g in trucks_df.groupby("priorityBucket"):
            sections[str(bucket)] = list(map(int, g["truckNumber"].tolist()))

    trucks_list = trucks_df.to_dict(
        orient="records") if not trucks_df.empty else []
    assigns_list = assigns_df.to_dict(
        orient="records") if not assigns_df.empty else []

    # Clean up NaN values in assignments before creating Pydantic models
    for assign in assigns_list:
        # Handle NaN values for boolean fields
        if pd.isna(assign.get('isRemainder')):
            assign['isRemainder'] = False
        if pd.isna(assign.get('isPartial')):
            assign['isPartial'] = False
        # Handle NaN values for string fields
        if pd.isna(assign.get('parentLine')):
            assign['parentLine'] = None
        if 'trttav_no' in assign and pd.isna(assign.get('trttav_no')):
            assign['trttav_no'] = None
        # Handle NaN values for integer fields
        if pd.isna(assign.get('remainingPieces')):
            assign['remainingPieces'] = 0

    # Ensure JSON-serializable keys and create typed models
    trucks_models: List[TruckSummary] = [TruckSummary(
        **{str(k): v for k, v in t.items()}) for t in trucks_list]
    assigns_models: List[LineAssignment] = [LineAssignment(
        **{str(k): v for k, v in a.items()}) for a in assigns_list]

    return OptimizeResponse(
        trucks=trucks_models,
        assignments=assigns_models,
        sections=sections,
    )


@app.on_event("startup")
def _startup_init() -> None:
    # Ensure geo/routing tables exist (graceful if DB not configured)
    _init_geo_tables()


@app.post("/geocode/validate")
async def geocode_validate(
    file: UploadFile = File(...),
    planningWhse: Optional[str] = Form("ZAC"),
) -> Dict[str, Any]:
    """Phase 1: extract addresses, geocode via Google if key present, and cache results.

    If GOOGLE_MAPS_API_KEY is not configured, returns detected addresses without lat/lng.
    Set planningWhse to empty string, "ALL", or "*" to process all warehouses.
    """
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="Only .xlsx files are supported")
    try:
        content: bytes = await file.read()
        buffer = BytesIO(content)
        df: pd.DataFrame = pd.read_excel(buffer, engine="openpyxl")
    except Exception as exc:
        raise HTTPException(
            status_code=400, detail=f"Failed to read Excel: {exc}") from exc

    # Filter by Planning Whse unless explicitly set to "all" or empty
    if planningWhse and planningWhse.upper() not in ("ALL", "*", ""):
        try:
            df = filter_by_planning_whse(df, allowed_values=(planningWhse,))
        except Exception:
            pass
    df = compute_calculated_fields(df)
    df = apply_routing_filters(df)

    addrs = extract_unique_addresses(df)

    api_key = os.getenv("GOOGLE_MAPS_API_KEY")
    dsn = _build_supabase_dsn()

    # BATCH CACHE LOOKUP - single query for all addresses
    cache_map: Dict[str, Dict[str, Any]] = {}
    if dsn:
        try:
            normalized_addrs = [a.get("normalized") or "" for a in addrs]
            with psycopg.connect(dsn, connect_timeout=30) as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        select normalized, latitude, longitude, confidence, provider
                        from address_cache
                        where normalized = ANY(%s)
                        """,
                        (normalized_addrs,),
                    )
                    for row in cur.fetchall():
                        cache_map[row[0]] = {
                            "latitude": row[1],
                            "longitude": row[2],
                            "confidence": row[3],
                            "provider": row[4],
                        }
            print(
                f"[geocode-validate] Batch cache lookup: {len(cache_map)} hits out of {len(addrs)} addresses")
        except Exception as exc:
            print(f"[geocode-validate] Batch cache lookup failed: {exc}")

    # Separate cached from uncached addresses
    enriched: List[Dict[str, Any]] = []
    cache_hits = 0
    api_calls = 0
    failures = 0
    new_geocodes = []  # For batch cache write

    # First pass: handle cache hits and identify cache misses
    cache_misses = []
    for a in addrs:
        normalized = a.get("normalized") or ""
        if normalized in cache_map:
            # Cache hit
            cached = cache_map[normalized]
            item = dict(a)
            item.update({
                "latitude": cached.get("latitude"),
                "longitude": cached.get("longitude"),
                "confidence": cached.get("confidence"),
                "provider": cached.get("provider"),
                "source": "cache",
            })
            enriched.append(item)
            cache_hits += 1
        else:
            cache_misses.append(a)

    # Parallel geocoding for cache misses
    if cache_misses and api_key:
        from concurrent.futures import ThreadPoolExecutor, as_completed

        def geocode_one(addr):
            """Geocode a single address and return result."""
            try:
                q = build_address_query(addr)
                lat, lng, conf, provider, _fmt = google_geocode_query(
                    q, api_key)
                return (addr, lat, lng, conf, provider, None)
            except Exception as exc:
                return (addr, None, None, None, None, str(exc))

        print(
            f"[geocode-validate] Geocoding {len(cache_misses)} addresses in parallel (max 10 concurrent)...")

        # Use ThreadPoolExecutor for parallel API calls (max 10 concurrent to be nice to API)
        with ThreadPoolExecutor(max_workers=10) as executor:
            future_to_addr = {executor.submit(
                geocode_one, a): a for a in cache_misses}

            for future in as_completed(future_to_addr):
                addr, lat, lng, conf, provider, error = future.result()
                item = dict(addr)
                normalized = addr.get("normalized") or ""

                if error:
                    item.update({"error": "GEOCODE_FAILED"})
                    failures += 1
                else:
                    item.update({
                        "latitude": lat,
                        "longitude": lng,
                        "confidence": conf,
                        "provider": provider,
                        "source": "google",
                    })
                    new_geocodes.append(
                        (normalized, addr, lat, lng, conf, provider))
                    api_calls += 1

                enriched.append(item)
    elif cache_misses and not api_key:
        # No API key configured
        for a in cache_misses:
            item = dict(a)
            item.update({"note": "No API key configured"})
            enriched.append(item)

    # BATCH CACHE WRITE - single transaction for all new geocodes
    if dsn and new_geocodes:
        try:
            with psycopg.connect(dsn, connect_timeout=30) as conn:
                with conn.cursor() as cur:
                    cur.executemany(
                        """
                        insert into address_cache (normalized, street, suite, city, state, zip, latitude, longitude, confidence, provider)
                        values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        on conflict (normalized)
                        do update set street=excluded.street, suite=excluded.suite, city=excluded.city,
                                      state=excluded.state, zip=excluded.zip, latitude=excluded.latitude,
                                      longitude=excluded.longitude, confidence=excluded.confidence,
                                      provider=excluded.provider, updated_at=now()
                        """,
                        [
                            (
                                normalized,
                                parts.get("street"),
                                parts.get("suite"),
                                parts.get("city"),
                                parts.get("state"),
                                parts.get("zip"),
                                lat,
                                lng,
                                conf,
                                provider,
                            )
                            for normalized, parts, lat, lng, conf, provider in new_geocodes
                        ],
                    )
                conn.commit()
            print(
                f"[geocode-validate] Batch cache write: {len(new_geocodes)} new geocodes")
        except Exception as exc:
            print(f"[geocode-validate] Batch cache write failed: {exc}")

    return {
        "count": len(enriched),
        "addresses": enriched,
        "cache_hits": cache_hits,
        "api_calls": api_calls,
        "failures": failures,
    }


@app.post("/distance-matrix")
async def distance_matrix(
    origins: str = Form(...),  # "lat,lng|lat,lng|..."
    destinations: str = Form(...),
) -> Dict[str, Any]:
    """Compute distance and duration matrices. Uses Google if key exists; fallback to Haversine.

    Caches results in distance_cache table to avoid repeated API calls.
    Accepts pipe-separated lat,lng pairs for origins and destinations.
    """
    def parse(s: str) -> List[tuple[float, float]]:
        out: List[tuple[float, float]] = []
        for tok in (s or "").split("|"):
            tok = tok.strip()
            if not tok:
                continue
            parts = tok.split(",")
            if len(parts) != 2:
                continue
            try:
                out.append((float(parts[0]), float(parts[1])))
            except Exception:
                continue
        return out

    def normalize_coord(lat: float, lng: float) -> str:
        """Normalize coordinate to 4 decimal places for cache key."""
        return f"{lat:.4f},{lng:.4f}"

    o = parse(origins)
    d = parse(destinations)
    if not o or not d:
        raise HTTPException(
            status_code=400, detail="Invalid origins/destinations")

    api_key = os.getenv("GOOGLE_MAPS_API_KEY")
    provider = "google" if api_key else "haversine"
    dsn = _build_supabase_dsn()

    # Check cache first
    dist: List[List[float]] = [
        [0.0 for _ in range(len(d))] for _ in range(len(o))]
    dur: List[List[float]] = [
        [0.0 for _ in range(len(d))] for _ in range(len(o))]
    cache_misses: List[tuple[int, int]] = []

    if dsn:
        try:
            with psycopg.connect(dsn, connect_timeout=5) as conn:
                with conn.cursor() as cur:
                    for i, orig in enumerate(o):
                        for j, dest in enumerate(d):
                            orig_key = normalize_coord(orig[0], orig[1])
                            dest_key = normalize_coord(dest[0], dest[1])
                            cur.execute(
                                """
                                select distance_miles, duration_minutes 
                                from distance_cache 
                                where origin_normalized = %s and dest_normalized = %s and provider = %s
                                """,
                                (orig_key, dest_key, provider)
                            )
                            row = cur.fetchone()
                            if row:
                                dist[i][j] = float(row[0] or 0.0)
                                dur[i][j] = float(row[1] or 0.0)
                            else:
                                cache_misses.append((i, j))
        except Exception as exc:
            print(f"[distance-matrix] Cache lookup failed: {exc}")
            # If cache fails, treat everything as cache miss
            cache_misses = [(i, j) for i in range(len(o))
                            for j in range(len(d))]
    else:
        # No database, compute everything
        cache_misses = [(i, j) for i in range(len(o)) for j in range(len(d))]

    # Compute cache misses
    if cache_misses:
        try:
            if api_key:
                try:
                    # Build origins/destinations lists for just the misses
                    # For efficiency, if we have many misses, just call the full matrix
                    if len(cache_misses) > len(o) * len(d) * 0.5:
                        # More than half are misses, compute full matrix
                        full_dist, full_dur = google_distance_matrix(
                            api_key, o, d)
                        for i in range(len(o)):
                            for j in range(len(d)):
                                dist[i][j] = full_dist[i][j]
                                dur[i][j] = full_dur[i][j]
                    else:
                        # Compute individual pairs (less efficient but saves API quota)
                        for i, j in cache_misses:
                            pair_dist, pair_dur = google_distance_matrix(
                                api_key, [o[i]], [d[j]])
                            dist[i][j] = pair_dist[0][0]
                            dur[i][j] = pair_dur[0][0]
                except Exception as gexc:
                    # Graceful fallback to Haversine if Google call fails
                    print(
                        f"[distance-matrix] Google failed; falling back to Haversine: {gexc}")
                    from .distance_service import haversine_matrix
                    coords = o + d
                    full_dist, full_dur = haversine_matrix(coords)
                    n_o = len(o)
                    n_d = len(d)
                    for i in range(n_o):
                        for j in range(n_d):
                            dist[i][j] = full_dist[i][j + n_o]
                            dur[i][j] = full_dur[i][j + n_o]
                    provider = "haversine"
            else:
                # No API key, use Haversine
                from .distance_service import haversine_matrix
                coords = o + d
                full_dist, full_dur = haversine_matrix(coords)
                n_o = len(o)
                n_d = len(d)
                for i in range(n_o):
                    for j in range(n_d):
                        dist[i][j] = full_dist[i][j + n_o]
                        dur[i][j] = full_dur[i][j + n_o]
        except Exception as exc:
            raise HTTPException(
                status_code=500, detail=f"Distance matrix failed: {exc}")

        # Cache the newly computed values
        if dsn and cache_misses:
            try:
                with psycopg.connect(dsn, connect_timeout=5) as conn:
                    with conn.cursor() as cur:
                        for i, j in cache_misses:
                            orig_key = normalize_coord(o[i][0], o[i][1])
                            dest_key = normalize_coord(d[j][0], d[j][1])
                            cur.execute(
                                """
                                insert into distance_cache (origin_normalized, dest_normalized, provider, distance_miles, duration_minutes)
                                values (%s, %s, %s, %s, %s)
                                on conflict (origin_normalized, dest_normalized, provider) 
                                do update set distance_miles = excluded.distance_miles, 
                                             duration_minutes = excluded.duration_minutes,
                                             updated_at = now()
                                """,
                                (orig_key, dest_key, provider,
                                 dist[i][j], dur[i][j])
                            )
                    conn.commit()
                print(
                    f"[distance-matrix] Cached {len(cache_misses)} new distance calculations")
            except Exception as exc:
                print(f"[distance-matrix] Cache write failed: {exc}")

    return {"distance_miles": dist, "duration_minutes": dur}


@app.get("/depot/location")
def depot_get() -> Dict[str, Any]:
    default_depot = {
        "id": 1,
        "name": "28",
        "address": "1155 NE 28th Street Fort Worth Tx 76106",
        "latitude": None,
        "longitude": None,
    }
    dsn = _build_supabase_dsn()
    if not dsn:
        return default_depot
    try:
        with psycopg.connect(dsn, connect_timeout=5) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "select id, name, address, latitude, longitude from depot_config where id=1")
                row = cur.fetchone()
                if not row:
                    return default_depot
                return {"id": row[0], "name": row[1], "address": row[2], "latitude": row[3], "longitude": row[4]}
    except Exception:
        return default_depot


@app.put("/depot/location")
def depot_put(name: Optional[str] = Form(None), address: Optional[str] = Form(None), latitude: Optional[float] = Form(None), longitude: Optional[float] = Form(None)) -> Dict[str, Any]:
    # If address is provided but no coordinates, try to geocode it
    if address and (latitude is None or longitude is None):
        api_key = os.getenv("GOOGLE_MAPS_API_KEY")
        if api_key:
            try:
                lat, lng, conf, provider, _fmt = google_geocode_query(
                    address, api_key)
                latitude = lat
                longitude = lng
                print(f"[depot:put] Geocoded depot address to {lat}, {lng}")
            except Exception as exc:
                print(f"[depot:put] Failed to geocode depot address: {exc}")

    # Default to Fort Worth if still no coordinates
    if latitude is None or latitude == 0:
        latitude = 32.795580
    if longitude is None or longitude == 0:
        longitude = -97.281410

    dsn = _build_supabase_dsn()
    if not dsn:
        # Accept but do not persist when DB is not configured
        return {"ok": True, "id": 1, "name": name, "address": address, "latitude": latitude, "longitude": longitude}
    try:
        with psycopg.connect(dsn, connect_timeout=5) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    insert into depot_config (id, name, address, latitude, longitude)
                    values (1, %s, %s, %s, %s)
                    on conflict (id) do update set name=excluded.name, address=excluded.address,
                        latitude=excluded.latitude, longitude=excluded.longitude, updated_at=now()
                    """,
                    (name, address, latitude, longitude),
                )
            conn.commit()
        return {"ok": True}
    except Exception as exc:
        # Graceful degradation per PRD: do not error if DB is unreachable/misconfigured
        # Return submitted values marked as not persisted
        print(f"[depot:put] Skipping persist; DB unavailable: {exc}")
        return {
            "ok": True,
            "id": 1,
            "name": name,
            "address": address,
            "latitude": latitude,
            "longitude": longitude,
            "note": "not persisted (DB unavailable)",
        }


# =========================
# Phase 2: Overweight pre-split then VRP remainder (weight-only)
# =========================

def _infer_country_from_state(state: str) -> str:
    s = str(state or "").strip().upper()
    MX = {
        "AGU", "BCN", "BCS", "CAM", "CHP", "CHH", "CH", "CMX", "COA", "COL",
        "DUR", "GUA", "GRO", "HID", "JAL", "MEX", "MIC", "MOR", "NAY", "NLE",
        "OAX", "PUE", "QUE", "ROO", "SLP", "SIN", "SON", "TAB", "TAM", "TLA",
        "VER", "YUC", "ZAC"
    }
    return "Mexico" if s in MX else "USA"


def _aggregate_sites(df: pd.DataFrame) -> List[Dict[str, Any]]:
    sites: List[Dict[str, Any]] = []
    if df.empty:
        return sites

    # Ensure necessary columns exist
    required_cols = ["Customer", "shipping_city", "shipping_state",
                     "RPcs", "Ready Weight"]
    for c in required_cols:
        if c not in df.columns:
            df[c] = None

    # Prepare per-piece weight
    if "Weight Per Piece" not in df.columns:
        df["Weight Per Piece"] = pd.to_numeric(
            df["Ready Weight"], errors="coerce") / df["RPcs"].replace(0, pd.NA)

    # Group by site (customer+city+state+country)
    tmp = df.copy()
    tmp["country"] = tmp["shipping_state"].apply(_infer_country_from_state)
    grp_cols = ["Customer", "shipping_city", "shipping_state", "country"]
    for (customer, city, state, country), g in tmp.groupby(grp_cols, dropna=False):
        total_weight = float(pd.to_numeric(
            g["Ready Weight"], errors="coerce").fillna(0).sum())
        total_pieces = int(pd.to_numeric(
            g["RPcs"], errors="coerce").fillna(0).sum())
        if total_weight <= 0 or total_pieces <= 0:
            continue
        site_id = f"{str(customer)}|{str(city)}|{str(state)}|{str(country)}"
        # capture line-level info used for splitting
        lines = []
        for _, r in g.iterrows():
            rp = int(pd.to_numeric(r.get("RPcs"), errors="coerce") or 0)
            rw = float(pd.to_numeric(
                r.get("Ready Weight"), errors="coerce") or 0.0)
            wpp = r.get("Weight Per Piece")
            try:
                wppf = float(wpp) if wpp is not None else (
                    rw / rp if rp > 0 else 0.0)
            except Exception:
                wppf = rw / rp if rp > 0 else 0.0
            if rp <= 0 or rw <= 0:
                continue
            lines.append({
                "SO": r.get("SO"),
                "Line": r.get("Line"),
                "RPcs": int(rp),
                "ReadyWeight": float(rw),
                "WeightPerPiece": float(wppf),
            })

        sites.append({
            "site_id": site_id,
            "customer": str(customer),
            "city": str(city),
            "state": str(state),
            "country": str(country),
            "total_weight": total_weight,
            "total_pieces": total_pieces,
            "latitude": None,
            "longitude": None,
            "lines": lines,
        })

    return sites


def _geocode_sites(sites: List[Dict[str, Any]]) -> None:
    api_key = os.getenv("GOOGLE_MAPS_API_KEY")
    dsn = _build_supabase_dsn()
    for s in sites:
        if s.get("latitude") and s.get("longitude"):
            continue
        # Use city/state/country as query
        parts = {"city": s.get("city"), "state": s.get(
            "state"), "country": s.get("country")}
        normalized = f"{str(parts.get('city') or '').strip().lower()},{str(parts.get('state') or '').strip().upper()},{str(parts.get('country') or '').strip()}"
        cached = cache_lookup_address(dsn, normalized) if dsn else None
        if cached and cached.get("latitude") and cached.get("longitude"):
            s["latitude"] = cached["latitude"]
            s["longitude"] = cached["longitude"]
            continue
        if not api_key:
            continue
        try:
            q = ", ".join(
                [c for c in [s.get("city"), s.get("state"), s.get("country")] if c])
            lat, lng, conf, provider, _ = google_geocode_query(q, api_key)
            s["latitude"], s["longitude"] = float(lat), float(lng)
            # Upsert into cache under our normalized key
            if dsn:
                cache_upsert_address(dsn, normalized, {"city": s.get("city"), "state": s.get(
                    "state"), "country": s.get("country")}, lat, lng, conf, provider)
        except Exception:
            continue


def _compute_depot_legs(depot_lat: float, depot_lng: float, sites: List[Dict[str, Any]]) -> Dict[str, Dict[str, float]]:
    # Build coords list for destinations
    dests = []
    idx_map: List[str] = []
    for s in sites:
        if s.get("latitude") and s.get("longitude"):
            dests.append((float(s["latitude"]), float(s["longitude"])))
            idx_map.append(s["site_id"])
    if not dests:
        return {}
    api_key = os.getenv("GOOGLE_MAPS_API_KEY")
    dist, dur = google_distance_matrix(
        api_key, origins=[(depot_lat, depot_lng)], destinations=dests)
    legs: Dict[str, Dict[str, float]] = {}
    # dist and dur are lists; first row corresponds to depot
    for j, site_id in enumerate(idx_map):
        miles = float(dist[0][j]) if dist and dist[0] else 0.0
        minutes = float(dur[0][j]) if dur and dur[0] else 0.0
        legs[site_id] = {"miles_out": miles, "minutes_out": minutes}
    return legs


def _presplit_overweight(sites: List[Dict[str, Any]], legs: Dict[str, Dict[str, float]], capacity: float, starting_truck_id: int = 1) -> (List[SingleStopTruck], List[Dict[str, Any]], int):
    trucks: List[SingleStopTruck] = []
    residual: List[Dict[str, Any]] = []
    next_id = starting_truck_id
    for s in sites:
        site_weight = float(s.get("total_weight") or 0.0)
        pieces_left = int(s.get("total_pieces") or 0)
        lines = list(s.get("lines") or [])
        # Sort by heavy pieces first
        lines = [{**ln} for ln in lines]
        lines.sort(key=lambda x: float(
            x.get("WeightPerPiece") or 0.0), reverse=True)

        if site_weight <= capacity + 1e-6:
            # No pre-split; pass along as residual
            residual.append(s)
            continue

        full_trucks = []
        remaining_weight = site_weight
        remaining_lines = lines

        while remaining_weight > capacity + 1e-6 and remaining_lines:
            take_weight = 0.0
            take_pcs = 0
            take_lines: List[Dict[str, Any]] = []
            new_remaining: List[Dict[str, Any]] = []
            for ln in remaining_lines:
                pcs = int(ln.get("RPcs") or 0)
                wpp = float(ln.get("WeightPerPiece") or 0.0)
                # take as many pieces as fit
                max_can_take = int((capacity - take_weight) //
                                   wpp) if wpp > 0 else 0
                if max_can_take <= 0:
                    new_remaining.append(ln)
                    continue
                take = min(pcs, max_can_take)
                if take > 0:
                    take_weight += take * wpp
                    take_pcs += take
                    left = pcs - take
                    if left > 0:
                        new_remaining.append(
                            {**ln, "RPcs": left, "ReadyWeight": left * wpp})
                    # record taken fragment (not stored now)
                else:
                    new_remaining.append(ln)
                if capacity - take_weight < 1e-6:
                    # truck is effectively full
                    # remaining lines stay for next trucks
                    new_remaining.extend(
                        [l for l in remaining_lines[remaining_lines.index(ln)+1:]])
                    break
            # finalize this truck if it has any weight
            if take_weight > 0:
                full_trucks.append({"weight": take_weight, "pieces": take_pcs})
                remaining_weight -= take_weight
                remaining_lines = new_remaining
            else:
                # no progress; break to avoid infinite loop
                break

        # Emit full single-stop trucks
        leg = legs.get(s["site_id"], {"miles_out": 0.0, "minutes_out": 0.0})
        for idx, ft in enumerate(full_trucks, 1):
            trucks.append(SingleStopTruck(
                truck_id=next_id,
                site_id=s["site_id"],
                customer=s["customer"],
                city=s["city"],
                state=s["state"],
                latitude=float(s.get("latitude") or 0.0),
                longitude=float(s.get("longitude") or 0.0),
                weight=float(ft["weight"]),
                pieces=int(ft["pieces"]),
                total_distance_miles=float(2 * leg.get("miles_out", 0.0)),
                total_duration_minutes=float(2 * leg.get("minutes_out", 0.0)),
            ))
            next_id += 1

        # Build residual site with remaining_lines aggregated
        res_weight = sum(float(ln.get("ReadyWeight") or (ln.get(
            "RPcs", 0) * float(ln.get("WeightPerPiece") or 0.0))) for ln in remaining_lines)
        res_pcs = sum(int(ln.get("RPcs") or 0) for ln in remaining_lines)
        if res_weight > 1e-6 and res_pcs > 0:
            residual.append({**s, "total_weight": float(res_weight),
                            "total_pieces": int(res_pcs), "lines": remaining_lines})

    return trucks, residual, next_id


@app.post("/route/v1/optimize")
async def route_optimize_v1(
    file: UploadFile = File(...),
    planningWhse: Optional[str] = Form("ZAC"),
    maxWeightPerTruck: Optional[int] = Form(52000),
    serviceTimePerStopMinutes: Optional[int] = Form(30),
    maxStopsPerRoute: Optional[int] = Form(20),
    maxTrucks: Optional[int] = Form(50),
) -> OptimizeV1Response:
    # 1) Read + filter + calculated fields
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="Only .xlsx files are supported")
    try:
        content: bytes = await file.read()
        df: pd.DataFrame = pd.read_excel(BytesIO(content), engine="openpyxl")
    except Exception as exc:
        raise HTTPException(
            status_code=400, detail=f"Failed to read Excel: {exc}") from exc

    if planningWhse:
        try:
            df = filter_by_planning_whse(df, allowed_values=(planningWhse,))
        except Exception:
            pass
    df = compute_calculated_fields(df)
    df = apply_routing_filters(df)

    # 2) Aggregate sites
    sites = _aggregate_sites(df)
    if not sites:
        raise HTTPException(status_code=400, detail="No valid sites found")

    # 3) Geocode sites
    _geocode_sites(sites)

    # 4) Depot
    depot = depot_get()
    depot_lat = depot.get("latitude") if depot.get("latitude") else 32.795580
    depot_lng = depot.get("longitude") if depot.get(
        "longitude") else -97.281410

    # 5) Depot legs O(N)
    legs = _compute_depot_legs(depot_lat, depot_lng, sites)

    # 6) Overweight pre-split (extract only full trucks; remainder stays for VRP)
    capacity = float(maxWeightPerTruck or 52000)
    single_trucks, residual_sites, next_truck_id = _presplit_overweight(
        sites, legs, capacity, starting_truck_id=1)

    # Filter residual to those <= capacity
    residual_sites = [s for s in residual_sites if float(
        s.get("total_weight") or 0.0) <= capacity + 1e-6]

    # 7) Build VRP for residual sites
    routes_vrp: List[Dict[str, Any]] = []
    if residual_sites:
        # Build coords including depot as index 0
        coords = [(depot_lat, depot_lng)] + [
            (float(s.get("latitude") or 0.0), float(s.get("longitude") or 0.0)) for s in residual_sites
        ]

        # Build matrices using Google with caching logic similar to phase2
        n = len(coords)
        dist_matrix = [[0.0] * n for _ in range(n)]
        dur_matrix = [[0.0] * n for _ in range(n)]

        api_key = os.getenv("GOOGLE_MAPS_API_KEY")
        dsn = _build_supabase_dsn()

        # For simplicity, fetch full matrix via one call when small; else pair-by-pair
        try:
            full_dist, full_dur = google_distance_matrix(
                api_key, coords, coords)
            for i in range(n):
                for j in range(n):
                    dist_matrix[i][j] = full_dist[i][j]
                    dur_matrix[i][j] = full_dur[i][j]
        except Exception as exc:
            # Fallback: compute pairs
            for i in range(n):
                for j in range(n):
                    if i == j:
                        continue
                    try:
                        pair_dist, pair_dur = google_distance_matrix(
                            api_key, [coords[i]], [coords[j]])
                        dist_matrix[i][j] = pair_dist[0][0]
                        dur_matrix[i][j] = pair_dur[0][0]
                    except Exception:
                        dist_matrix[i][j] = 0.0
                        dur_matrix[i][j] = 0.0

        # Build stops
        stops: List[Stop] = []
        for s in residual_sites:
            stops.append(Stop(
                customer_name=str(s.get("customer")),
                address=f"{s.get('city')}, {s.get('state')}",
                city=str(s.get("city")),
                state=str(s.get("state")),
                latitude=float(s.get("latitude") or 0.0),
                longitude=float(s.get("longitude") or 0.0),
                weight=float(s.get("total_weight") or 0.0),
                pieces=int(s.get("total_pieces") or 0),
                order_id=str(s.get("site_id")),
                line_id="site",
            ))

        vrp_routes = solve_vrp(
            stops=stops,
            depot_lat=depot_lat,
            depot_lng=depot_lng,
            distance_matrix=dist_matrix,
            duration_matrix=dur_matrix,
            max_weight_per_truck=capacity,
            max_drive_time_minutes=720,
            service_time_per_stop_minutes=float(
                serviceTimePerStopMinutes or 30),
            max_vehicles=int(maxTrucks or 50),
        )

        for r in vrp_routes:
            rd = r.to_dict()
            rd["route_type"] = "vrp"
            routes_vrp.append(rd)

    # 8) Assemble final
    routes_single = [t.model_dump() for t in single_trucks]
    all_routes = routes_single + routes_vrp
    totals = {
        "trucks": len(all_routes),
        "stops": sum(len(r.get("stops", [1])) if r.get("route_type") == "vrp" else 1 for r in all_routes),
        "weight": sum(
            (r.get("total_weight") or r.get("weight") or 0.0) for r in all_routes
        ),
    }
    diagnostics = {
        "overweight_sites": len([s for s in sites if float(s.get("total_weight") or 0.0) > capacity + 1e-6]),
        "residual_sites": len(residual_sites),
    }

    return OptimizeV1Response(routes=all_routes, depot={"latitude": depot_lat, "longitude": depot_lng, "name": depot.get("name")}, totals=totals, diagnostics=diagnostics)


@app.post("/route/optimize-phase2")
async def route_optimize_phase2(
    file: UploadFile = File(...),
    planningWhse: Optional[str] = Form("ZAC"),
    maxWeightPerTruck: Optional[int] = Form(52000),
    maxStopsPerTruck: Optional[int] = Form(20),
    maxDriveTimeMinutes: Optional[int] = Form(720),  # 12 hours default
    serviceTimePerStopMinutes: Optional[int] = Form(30),  # 30 min per stop
    maxTrucks: Optional[int] = Form(50),  # Max vehicles to use
) -> Dict[str, Any]:
    """Phase 2: Geographic clustering + TSP route optimization.

    Returns optimized routes with stop sequences and map visualization data.
    """
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="Only .xlsx files are supported")

    try:
        content: bytes = await file.read()
        buffer = BytesIO(content)
        df: pd.DataFrame = pd.read_excel(buffer, engine="openpyxl")
    except Exception as exc:
        raise HTTPException(
            status_code=400, detail=f"Failed to read Excel: {exc}") from exc

    # Filter by planning warehouse
    if planningWhse:
        try:
            df = filter_by_planning_whse(df, allowed_values=(planningWhse,))
        except Exception:
            pass

    df = compute_calculated_fields(df)
    df = apply_routing_filters(df)

    print(f"[route-optimize] Starting with {len(df)} filtered order lines")

    # Extract and geocode addresses
    addrs = extract_unique_addresses(df)
    api_key = os.getenv("GOOGLE_MAPS_API_KEY")
    dsn = _build_supabase_dsn()

    # Geocode all addresses
    geocoded_addrs: List[Dict[str, Any]] = []
    for a in addrs:
        cached = cache_lookup_address(dsn, a.get("normalized") or "")
        if cached and cached.get("latitude") and cached.get("longitude"):
            geocoded_addrs.append({
                **a,
                "latitude": cached["latitude"],
                "longitude": cached["longitude"],
                "confidence": cached.get("confidence", 0.0),
            })
        elif api_key:
            try:
                q = build_address_query(a)
                lat, lng, conf, provider, _fmt = google_geocode_query(
                    q, api_key)
                geocoded_addrs.append({
                    **a,
                    "latitude": lat,
                    "longitude": lng,
                    "confidence": conf,
                })
                cache_upsert_address(dsn, a.get(
                    "normalized") or "", a, lat, lng, conf, provider)
            except Exception:
                continue

    if not geocoded_addrs:
        raise HTTPException(
            status_code=400, detail="No addresses could be geocoded")

    # Get depot location
    depot = depot_get()
    depot_lat = depot.get("latitude") if depot.get("latitude") else 32.795580
    depot_lng = depot.get("longitude") if depot.get(
        "longitude") else -97.281410

    # Build distance matrix (depot + all geocoded addresses) WITH CACHING
    coords = [(depot_lat, depot_lng)] + [(a["latitude"], a["longitude"])
                                         for a in geocoded_addrs]

    # Initialize result matrices
    n = len(coords)
    dist_matrix = [[0.0] * n for _ in range(n)]
    dur_matrix = [[0.0] * n for _ in range(n)]

    # Check cache for all pairs in a single batch query (much faster!)
    cache_misses = []
    provider = "cached"

    if dsn:
        # Build all coordinate keys
        coord_keys = [
            f"{coords[i][0]:.6f},{coords[i][1]:.6f}" for i in range(n)]

        # Build list of all pairs we need
        all_pairs = []
        for i in range(n):
            for j in range(n):
                if i != j:
                    all_pairs.append((i, j, coord_keys[i], coord_keys[j]))

        try:
            # Single batch query for all cache lookups
            with psycopg.connect(dsn, connect_timeout=30) as conn:
                with conn.cursor() as cur:
                    # Use unnest to query all pairs at once
                    pairs_tuples = [(orig, dest)
                                    for _, _, orig, dest in all_pairs]

                    cur.execute(
                        """
                        select origin_normalized, dest_normalized, distance_miles, duration_minutes
                        from distance_cache
                        where (origin_normalized, dest_normalized) = ANY(%s)
                        and provider = 'google'
                        """,
                        (pairs_tuples,),
                    )

                    # Build cache hit map
                    cache_hits = {}
                    for row in cur.fetchall():
                        key = (row[0], row[1])  # (origin, dest)
                        cache_hits[key] = (float(row[2]), float(
                            row[3]))  # (distance, duration)

                    print(
                        f"[route-optimize] Found {len(cache_hits)} cached pairs out of {len(all_pairs)}")

                    # Fill in matrix from cache and track misses
                    for i, j, orig_key, dest_key in all_pairs:
                        key = (orig_key, dest_key)
                        if key in cache_hits:
                            dist_matrix[i][j] = cache_hits[key][0]
                            dur_matrix[i][j] = cache_hits[key][1]
                        else:
                            cache_misses.append((i, j))
        except Exception as exc:
            print(
                f"[route-optimize] Batch cache lookup failed: {exc}, treating all as cache misses")
            cache_misses = [(i, j) for i, j, _, _ in all_pairs]
    else:
        # No DB connection, all are cache misses
        for i in range(n):
            for j in range(n):
                if i != j:
                    cache_misses.append((i, j))

    # Fetch missing pairs from Google API or fallback to Haversine
    if cache_misses and api_key:
        print(
            f"[route-optimize] Cache misses: {len(cache_misses)} out of {n*n} pairs")

        # For large matrices (>100 addresses), Google API is too slow/expensive
        # Use Haversine as fallback for first run, then cache for future runs
        if n > 100:
            print(
                f"[route-optimize] Large dataset ({n} addresses), using Haversine distance approximation")
            print("[route-optimize] Results will be cached for future use")
            haversine_dist, haversine_dur = haversine_matrix(coords)
            for i, j in cache_misses:
                dist_matrix[i][j] = haversine_dist[i][j]
                dur_matrix[i][j] = haversine_dur[i][j]
            provider = "haversine"
        else:
            try:
                # For smaller datasets, use Google API
                if len(cache_misses) > n * n * 0.5:
                    # Most are misses, compute full matrix
                    full_dist, full_dur = google_distance_matrix(
                        api_key, coords, coords)
                    for i in range(n):
                        for j in range(n):
                            dist_matrix[i][j] = full_dist[i][j]
                            dur_matrix[i][j] = full_dur[i][j]
                else:
                    # Compute individual pairs (fewer misses)
                    for i, j in cache_misses:
                        pair_dist, pair_dur = google_distance_matrix(
                            api_key, [coords[i]], [coords[j]])
                        dist_matrix[i][j] = pair_dist[0][0]
                        dur_matrix[i][j] = pair_dur[0][0]

                provider = "google"
            except Exception as gexc:
                print(
                    f"[route-optimize] Google failed; falling back to Haversine: {gexc}")
                haversine_dist, haversine_dur = haversine_matrix(coords)
                for i, j in cache_misses:
                    dist_matrix[i][j] = haversine_dist[i][j]
                    dur_matrix[i][j] = haversine_dur[i][j]
                provider = "haversine"

        # Write to cache in batch
        if dsn and cache_misses:
            try:
                with psycopg.connect(dsn, connect_timeout=30) as conn:
                    with conn.cursor() as cur:
                        # Batch insert using executemany
                        cache_data = []
                        for i, j in cache_misses:
                            orig_key = f"{coords[i][0]:.6f},{coords[i][1]:.6f}"
                            dest_key = f"{coords[j][0]:.6f},{coords[j][1]:.6f}"
                            cache_data.append((
                                orig_key, dest_key, provider,
                                dist_matrix[i][j], dur_matrix[i][j]
                            ))

                        cur.executemany(
                            """insert into distance_cache (origin_normalized, dest_normalized, provider, distance_miles, duration_minutes)
                               values (%s,%s,%s,%s,%s)
                               on conflict (origin_normalized, dest_normalized, provider)
                               do update set distance_miles = excluded.distance_miles, 
                                           duration_minutes = excluded.duration_minutes""",
                            cache_data
                        )
                    conn.commit()
                    print(
                        f"[route-optimize] Cached {len(cache_misses)} new distance calculations")
            except Exception as exc:
                print(f"[route-optimize] Cache write failed: {exc}")

    elif cache_misses and not api_key:
        # No API key, use Haversine for all misses
        haversine_dist, haversine_dur = haversine_matrix(coords)
        for i, j in cache_misses:
            dist_matrix[i][j] = haversine_dist[i][j]
            dur_matrix[i][j] = haversine_dur[i][j]
    else:
        print(f"[route-optimize] All {n*n} pairs loaded from cache!")

    # Create Stop objects for route optimization
    # Strategy: Pack customer orders into truck-sized loads (bin packing by customer)
    stops: List[Stop] = []
    unroutable_lines: List[Dict[str, Any]] = []

    max_weight = float(maxWeightPerTruck or 52000)

    # Prepare geocoded address lookup
    addr_lookup = {}
    for addr in geocoded_addrs:
        city_key = str(addr.get("city", "")).lower()
        state_key = str(addr.get("state", "")).upper()
        addr_lookup[(city_key, state_key)] = addr

    # Debug file write to confirm this code is running
    from pathlib import Path
    debug_file = Path(__file__).parent.parent / "LAST_RUN.txt"
    with open(debug_file, "w") as f:
        f.write(f"Bin packing started at {_dt.datetime.now()}\n")
        f.write(f"Order lines: {len(df)}\n")
        f.write(f"Customers: {df['Customer'].nunique()}\n")
        f.write(f"Max weight: {max_weight:,.0f} lbs\n")

    print(
        f"[route-optimize] Starting truck packing with max weight: {max_weight:,.0f} lbs")
    print(
        f"[route-optimize] Processing {len(df)} order lines for {df['Customer'].nunique()} customers")

    # Group by Customer + Destination for bin packing
    for (customer, city, state), customer_df in df.groupby(["Customer", "shipping_city", "shipping_state"], dropna=False):
        # Get geocoded location for this customer/destination
        city_key = str(city).lower()
        state_key = str(state).upper()
        addr = addr_lookup.get((city_key, state_key))

        if not addr or not addr.get("latitude") or not addr.get("longitude"):
            print(
                f"[route-optimize] Skipping {customer} -> {city}, {state} (no geocode)")
            continue

        # Sort lines by weight (descending) for better bin packing
        customer_lines = customer_df.sort_values(
            "Ready Weight", ascending=False).to_dict('records')

        total_customer_weight = sum(line["Ready Weight"]
                                    for line in customer_lines)
        print(
            f"[route-optimize] Packing {customer} -> {city}, {state}: {total_customer_weight:,.0f} lbs ({len(customer_lines)} lines)")

        # Bin packing algorithm: piece-level packing to ensure each load <= max_weight
        truck_loads: List[Dict[str, Any]] = []
        current_truck: Dict[str, Any] = {
            "weight": 0.0, "pieces": 0, "lines": []}

        for line in customer_lines:
            # Extract counts and weights
            line_pieces_total = int(line.get("RPcs") or 0)
            if line_pieces_total <= 0:
                continue

            weight_per_piece_raw = line.get("Weight Per Piece")
            try:
                weight_per_piece = float(
                    weight_per_piece_raw) if weight_per_piece_raw is not None else None
            except Exception:
                weight_per_piece = None

            if not weight_per_piece or weight_per_piece <= 0:
                # Fallback from total weight
                try:
                    lw = float(line.get("Ready Weight") or 0.0)
                except Exception:
                    lw = 0.0
                weight_per_piece = (
                    lw / line_pieces_total) if line_pieces_total > 0 else 0.0

            # If a single piece exceeds max, mark unroutable and skip
            if weight_per_piece > max_weight:
                unroutable_lines.append({
                    "SO": line.get("SO"),
                    "Line": line.get("Line"),
                    "Customer": line.get("Customer"),
                    "shipping_city": line.get("shipping_city"),
                    "shipping_state": line.get("shipping_state"),
                    "weight_per_piece": float(weight_per_piece),
                    "reason": "piece_weight_exceeds_truck_capacity",
                })
                continue

            pieces_remaining = line_pieces_total
            # Allocate pieces across trucks while respecting capacity
            while pieces_remaining > 0:
                available_capacity = max_weight - \
                    float(current_truck["weight"])
                # If nothing fits, finalize current truck and start a new one
                if available_capacity < weight_per_piece - 1e-6:
                    if current_truck["lines"]:
                        truck_loads.append(current_truck)
                    current_truck = {"weight": 0.0, "pieces": 0, "lines": []}
                    available_capacity = max_weight

                max_pieces_fit = int(available_capacity // weight_per_piece)
                if max_pieces_fit <= 0:
                    # Defensive: start a new truck if somehow we still can't fit a piece
                    if current_truck["lines"]:
                        truck_loads.append(current_truck)
                    current_truck = {"weight": 0.0, "pieces": 0, "lines": []}
                    continue

                take = min(pieces_remaining, max_pieces_fit)
                # Create a fragment line representing the allocated pieces
                frag_weight = take * weight_per_piece
                line_fragment = dict(line)
                line_fragment["RPcs"] = int(take)
                line_fragment["Ready Weight"] = float(frag_weight)

                current_truck["weight"] = float(
                    current_truck["weight"]) + float(frag_weight)
                current_truck["pieces"] = int(
                    current_truck["pieces"]) + int(take)
                current_truck["lines"].append(line_fragment)
                pieces_remaining -= take

        # Don't forget the last truck
        if current_truck["lines"]:
            truck_loads.append(current_truck)

        # Create a Stop for each truck load
        for truck_idx, truck in enumerate(truck_loads, 1):
            so_numbers = [str(line.get("SO", "")) for line in truck["lines"]]
            order_id = f"{customer[:20]}-T{truck_idx}" if len(
                truck_loads) > 1 else str(customer[:30])

            stops.append(Stop(
                customer_name=str(customer),
                address=addr.get("street", "") or addr.get("normalized", ""),
                city=str(city),
                state=str(state),
                latitude=addr["latitude"],
                longitude=addr["longitude"],
                weight=truck["weight"],
                pieces=truck["pieces"],
                order_id=order_id,
                # Indicate how many lines in this load
                line_id=f"{len(truck['lines'])} lines",
            ))

        if len(truck_loads) > 1:
            print(f"  -> Split into {len(truck_loads)} truck loads")

    print(
        f"[route-optimize] Created {len(stops)} truck loads from {len(df)} order lines")

    # Update debug file with detailed results
    with open(debug_file, "w") as f:
        f.write(f"[/route/optimize-phase2] Started at {_dt.datetime.now()}\n")
        f.write(f"Order lines after filters: {len(df)}\n")
        f.write(f"Max weight per truck: {max_weight:,.0f} lbs\n\n")
        f.write(f"Bin packing completed:\n")
        f.write(f"Truck loads created: {len(stops)}\n")
        f.write(f"Total weight: {sum(s.weight for s in stops):,.0f} lbs\n")
        f.write(
            f"Avg weight per load: {sum(s.weight for s in stops) / len(stops):,.0f} lbs\n\n")
        if unroutable_lines:
            f.write(
                f"Unroutable lines (piece overweight): {len(unroutable_lines)}\n")
            for ur in unroutable_lines[:200]:
                # Limit listing to avoid giant files
                f.write(
                    f"  SO {ur.get('SO')} Line {ur.get('Line')} - {ur.get('Customer')} {ur.get('shipping_city')}, {ur.get('shipping_state')} | WPP {ur.get('weight_per_piece'):,.0f}\n"
                )
            if len(unroutable_lines) > 200:
                f.write(f"  ... and {len(unroutable_lines) - 200} more\n")
        f.write(f"Stop details:\n")
        for i, stop in enumerate(stops):
            f.write(
                f"  {i}: {stop.customer_name[:30]}, {stop.city} - {stop.weight:,.0f} lbs, {stop.pieces} pcs\n")

    if not stops:
        raise HTTPException(status_code=400, detail="No valid stops found")

    # Plan routes using OR-Tools VRP solver
    routes = solve_vrp(
        stops=stops,
        depot_lat=depot_lat,
        depot_lng=depot_lng,
        distance_matrix=dist_matrix,
        duration_matrix=dur_matrix,
        max_weight_per_truck=float(maxWeightPerTruck or 52000),
        max_drive_time_minutes=float(maxDriveTimeMinutes or 720),
        service_time_per_stop_minutes=float(serviceTimePerStopMinutes or 30),
        max_vehicles=int(maxTrucks or 50),
    )

    # Log VRP results and compute drop reasons
    with open(debug_file, "a") as f:
        f.write(f"\nVRP Results:\n")
        f.write(f"Routes generated: {len(routes)}\n")
        routed_stop_indices = set()
        for route in routes:
            for stop in route.stops:
                # Find the index of this stop in the original stops list
                for i, s in enumerate(stops):
                    if s.order_id == stop.order_id:
                        routed_stop_indices.add(i)
                        break
        dropped_indices = [i for i in range(
            len(stops)) if i not in routed_stop_indices]
        f.write(f"Stops routed: {len(routed_stop_indices)} / {len(stops)}\n")
        f.write(f"Stops DROPPED by VRP: {len(dropped_indices)}\n")

        # Classify drop reasons per stop
        drop_reasons: list[dict[str, Any]] = []

        def classify_reason(i: int) -> dict[str, Any]:
            st = stops[i]
            # Roundtrip time including service time at the stop
            try:
                out_min = float(dur_matrix[0][i + 1])
                back_min = float(dur_matrix[i + 1][0])
            except Exception:
                out_min = 0.0
                back_min = 0.0
            service_min = float(serviceTimePerStopMinutes or 30)
            rt_minutes = out_min + back_min + service_min
            limit_minutes = float(maxDriveTimeMinutes or 720)
            capacity_limit = float(maxWeightPerTruck or 52000)

            if st.weight > capacity_limit + 1e-6:
                reason = "stop_weight_exceeds_truck_capacity"
            elif rt_minutes > limit_minutes + 1e-6:
                reason = "roundtrip_time_exceeds_limit"
            elif (out_min <= 0 and back_min <= 0):
                reason = "no_distance_time_available"
            else:
                reason = "not_routed_under_constraints"
            return {
                "index": i,
                "customer_name": st.customer_name,
                "city": st.city,
                "state": st.state,
                "weight": float(st.weight),
                "roundtrip_minutes": float(rt_minutes),
                "limit_minutes": float(limit_minutes),
                "reason": reason,
            }

        if dropped_indices:
            for idx in dropped_indices:
                drop_reasons.append(classify_reason(idx))
            f.write(f"Dropped stop indices: {dropped_indices}\n")
            f.write(f"Dropped stops with reasons:\n")
            for dr in drop_reasons:
                f.write(
                    f"  {dr['index']}: {dr['customer_name'][:30]}, {dr['city']} - {dr['weight']:,.0f} lbs | reason={dr['reason']} | rt_min={dr['roundtrip_minutes']:.1f} / limit={dr['limit_minutes']:.0f}\n"
                )

    return {
        "success": True,
        "routes": [r.to_dict() for r in routes],
        "depot": {"latitude": depot_lat, "longitude": depot_lng, "name": depot.get("name")},
        "total_trucks": len(routes),
        "total_stops": len(stops),
        "unroutable": unroutable_lines,
        # Diagnostic: reasons for dropped stops
        "drop_reasons": drop_reasons if 'drop_reasons' in locals() else [],
    }


@app.post("/route/plan", response_model=OptimizeResponse)
async def route_plan(
    file: UploadFile = File(...),
    planningWhse: Optional[str] = Form("ZAC"),
    deliveryDate: Optional[str] = Form(None),
    startLocation: Optional[str] = Form(None),
    endLocation: Optional[str] = Form(None),
    truckHours: Optional[int] = Form(10),
    minutesPerStop: Optional[int] = Form(30),
    texasMaxWeight: Optional[int] = Form(52000),
    otherMaxWeight: Optional[int] = Form(48000),
) -> OptimizeResponse:
    """Phase 1 routing foundation endpoint.

    For now, this reuses deterministic grouping while honoring:
    - Planning Whse filter
    - Optional delivery date filter against Earliest Due (<= deliveryDate)
    - Max weight limits per TX vs other states

    Future phases will compute geocoding, distance matrices, and route ordering.
    """
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="Only .xlsx files are supported")

    try:
        content: bytes = await file.read()
        buffer = BytesIO(content)
        df: pd.DataFrame = pd.read_excel(buffer, engine="openpyxl")
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=400, detail=f"Failed to read Excel: {exc}") from exc

    # Align with Upload/Optimize filters first
    if planningWhse:
        try:
            df = filter_by_planning_whse(df, allowed_values=(planningWhse,))
        except Exception:
            pass

    df = compute_calculated_fields(df)
    df = apply_routing_filters(df)

    # Optional delivery date filter: include rows whose Earliest Due <= deliveryDate
    if deliveryDate:
        try:
            cutoff = pd.to_datetime(deliveryDate, errors="coerce")
        except Exception:
            cutoff = pd.NaT
        if pd.notna(cutoff):
            # Try common header names
            candidates = [
                "Earliest Due",
                "earliest_due",
                "earliestDue",
            ]
            use_col = None
            for c in candidates:
                if c in df.columns:
                    use_col = c
                    break
            if use_col is None:
                # Fallback: try to coerce any 'Earliest' like column
                for c in df.columns:
                    if str(c).strip().lower().replace(" ", "").startswith("earliest"):
                        use_col = c
                        break
            if use_col is not None:
                try:
                    ed = pd.to_datetime(df[use_col], errors="coerce")
                    df = df.loc[(pd.isna(ed)) | (ed <= cutoff)].copy()
                except Exception:
                    pass

    cfg = {
        "texas_max_lbs": int(texasMaxWeight or 52000),
        "texas_min_lbs": 47000,  # keep existing min defaults for now
        "other_max_lbs": int(otherMaxWeight or 48000),
        "other_min_lbs": 44000,
    }

    try:
        trucks_df, assigns_df = naive_grouping(df, cfg)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=400, detail=f"Routing plan failed: {exc}") from exc

    sections: Dict[str, List[int]] = {}
    if not trucks_df.empty:
        for bucket, g in trucks_df.groupby("priorityBucket"):
            sections[str(bucket)] = list(map(int, g["truckNumber"].tolist()))

    trucks_list = trucks_df.to_dict(
        orient="records") if not trucks_df.empty else []
    assigns_list = assigns_df.to_dict(
        orient="records") if not assigns_df.empty else []

    # Clean NaNs like in /optimize
    for assign in assigns_list:
        if pd.isna(assign.get('isRemainder')):
            assign['isRemainder'] = False
        if pd.isna(assign.get('isPartial')):
            assign['isPartial'] = False
        if pd.isna(assign.get('parentLine')):
            assign['parentLine'] = None
        if 'trttav_no' in assign and pd.isna(assign.get('trttav_no')):
            assign['trttav_no'] = None
        if pd.isna(assign.get('remainingPieces')):
            assign['remainingPieces'] = 0

    trucks_models: List[TruckSummary] = [TruckSummary(
        **{str(k): v for k, v in t.items()}) for t in trucks_list]
    assigns_models: List[LineAssignment] = [LineAssignment(
        **{str(k): v for k, v in a.items()}) for a in assigns_list]

    return OptimizeResponse(
        trucks=trucks_models,
        assignments=assigns_models,
        sections=sections,
    )


@app.post("/export/trucks")
async def export_trucks(
    file: UploadFile = File(...),
    planningWhse: Optional[str] = Form("ZAC"),
) -> StreamingResponse:
    """Export optimized truck assignments to Excel format"""
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="Only .xlsx files are supported")

    try:
        content: bytes = await file.read()
        buffer = BytesIO(content)
        df: pd.DataFrame = pd.read_excel(buffer, engine="openpyxl")
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=400, detail=f"Failed to read Excel: {exc}") from exc

    # Apply Planning Whse filter first (default ZAC); if missing column, no-op
    if planningWhse:
        try:
            df = filter_by_planning_whse(df, allowed_values=(planningWhse,))
        except Exception:
            pass
    df = compute_calculated_fields(df)
    cfg = {
        "texas_max_lbs": 52000,
        "texas_min_lbs": 47000,
        "other_max_lbs": 48000,
        "other_min_lbs": 44000,
    }

    try:
        trucks_df, assigns_df = naive_grouping(df, cfg)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=400, detail=f"Optimization failed: {exc}") from exc

    # Create Excel output with multiple sheets
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if not trucks_df.empty:
            trucks_df.to_excel(writer, sheet_name="Truck Summary", index=False)
        if not assigns_df.empty:
            assigns_df.to_excel(
                writer, sheet_name="Order Details", index=False)

    output.seek(0)

    return StreamingResponse(
        BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=truck_optimization_results.xlsx"}
    )


@app.get("/no-multi-stop-customers")
def get_no_multi_stop_customers() -> Dict[str, List[str]]:
    """Get list of customers that cannot be combined with others"""
    return {"customers": NO_MULTI_STOP_CUSTOMERS}


@app.post("/no-multi-stop-customers")
def update_no_multi_stop_customers(customers: List[str]) -> Dict[str, str]:
    """Update list of customers that cannot be combined with others"""
    global NO_MULTI_STOP_CUSTOMERS
    # Note: This is temporary - in production this should be stored in a database
    NO_MULTI_STOP_CUSTOMERS.clear()
    NO_MULTI_STOP_CUSTOMERS.extend(customers)
    return {"message": f"Updated no-multi-stop list with {len(customers)} customers"}


@app.post("/combine-trucks", response_model=CombineTrucksResponse)
async def combine_trucks(
    file: UploadFile = File(...),
    request: str = Form(...),
    planningWhse: Optional[str] = Form(None),
) -> CombineTrucksResponse:
    """Combine selected lines into a single truck.

    Notes:
    - Stateless: recomputes trucks/assignments from the uploaded file.
    - Matches selected lines by (SO, Line) ignoring the client-side truck numbers to avoid numbering drift.
    - Chooses the lightest truck among those containing the selected lines as the target truck.
    """
    # Parse request payload
    try:
        req = CombineTrucksRequest.model_validate_json(request)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=400, detail=f"Invalid request: {exc}") from exc

    # Load Excel
    try:
        content: bytes = await file.read()
        buffer = BytesIO(content)
        df: pd.DataFrame = pd.read_excel(buffer, engine="openpyxl")
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=400, detail=f"Failed to read Excel: {exc}") from exc

    # Compute fields and optimize (deterministic grouping)
    # Optional: filter to Planning Whse to align with UI selection
    if planningWhse:
        try:
            df = filter_by_planning_whse(df, allowed_values=(planningWhse,))
        except Exception:
            pass
    df = compute_calculated_fields(df)
    df = apply_routing_filters(df)
    cfg = {
        "texas_max_lbs": req.weightConfig.texas_max_lbs,
        "texas_min_lbs": req.weightConfig.texas_min_lbs,
        "other_max_lbs": req.weightConfig.other_max_lbs,
        "other_min_lbs": req.weightConfig.other_min_lbs,
    }
    try:
        trucks_df, assigns_df = naive_grouping(df.copy(), cfg)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=400, detail=f"Optimization failed: {exc}") from exc

    if trucks_df.empty or assigns_df.empty:
        return CombineTrucksResponse(success=False, message="No trucks or assignments to combine", updatedAssignments=[], removedTruckIds=[])

    # Parse selected lines: id format "truckNumber-SO-Line"
    sel_pairs: List[tuple[str, str]] = []
    for lid in req.lineIds:
        parts = str(lid).split("-")
        if len(parts) < 3:
            continue
        _, so, line = parts[-3], parts[-2], parts[-1]
        sel_pairs.append((str(so), str(line)))
    if not sel_pairs:
        return CombineTrucksResponse(success=False, message="No valid line IDs provided", updatedAssignments=[], removedTruckIds=[])

    # Find matching assignment rows
    def as_str(s: Any) -> str:
        return str(s)

    assigns_df = assigns_df.copy()
    assigns_df["__so_key__"] = assigns_df["so"].map(as_str)
    assigns_df["__line_key__"] = assigns_df["line"].map(as_str)
    mask_selected = assigns_df.apply(lambda r: (
        r["__so_key__"], r["__line_key__"]) in sel_pairs, axis=1)
    selected_rows = assigns_df[mask_selected]
    if selected_rows.empty:
        return CombineTrucksResponse(success=False, message="Selected lines not found in current optimization", updatedAssignments=[], removedTruckIds=[])

    # Determine candidate trucks and target (lightest among involved)
    involved_trucks = sorted(set(int(t)
                             for t in selected_rows["truckNumber"].tolist()))
    trucks_sub = trucks_df[trucks_df["truckNumber"].isin(involved_trucks)]
    if trucks_sub.empty:
        return CombineTrucksResponse(success=False, message="Candidate trucks not found", updatedAssignments=[], removedTruckIds=[])
    target_row = trucks_sub.sort_values("totalWeight").iloc[0]
    target_truck = int(target_row["truckNumber"])  # lightest

    # Compute total selected weight and validate against target's max
    total_selected_weight = float(selected_rows["totalWeight"].sum())
    target_max = float(target_row.get("maxWeight") or 0.0)
    if target_max > 0 and (float(target_row.get("totalWeight") or 0.0) + total_selected_weight) > target_max * 1.0001:
        return CombineTrucksResponse(success=False, message="Combination exceeds target truck max weight", updatedAssignments=[], removedTruckIds=[])

    # Reassign selected lines to target
    source_trucks_before = set(involved_trucks)
    assigns_df.loc[mask_selected, "truckNumber"] = target_truck

    # Determine removed trucks (those that had only selected lines)
    remaining_by_truck = assigns_df.groupby("truckNumber").size()
    removed = [t for t in source_trucks_before if t !=
               target_truck and t not in remaining_by_truck.index.tolist()]

    # Recompute target truck summary from its assignments
    t_assigns = assigns_df[assigns_df["truckNumber"] == target_truck].copy()
    # Determine state and limits from any row
    any_state = str(t_assigns["customerState"].iloc[0]
                    ) if not t_assigns.empty else ""
    is_texas = str(any_state).strip().upper() in {"TX", "TEXAS"}
    max_weight = cfg["texas_max_lbs"] if is_texas else cfg["other_max_lbs"]
    min_weight = cfg["texas_min_lbs"] if is_texas else cfg["other_min_lbs"]
    total_weight = float(t_assigns["totalWeight"].sum())
    total_pieces = int(t_assigns["piecesOnTransport"].sum())
    total_lines = int(t_assigns.shape[0])
    total_orders = int(t_assigns["so"].nunique())
    max_width = float(t_assigns["width"].max()) if not t_assigns.empty else 0.0
    contains_late = bool(t_assigns["isLate"].any(
    )) if "isLate" in t_assigns.columns else False
    priority_bucket = "Late" if contains_late else "WithinWindow"
    # Use first destination/customer for summary context
    customer_name = str(
        t_assigns["customerName"].iloc[0]) if not t_assigns.empty else ""
    customer_city = str(
        t_assigns["customerCity"].iloc[0]) if not t_assigns.empty else ""
    customer_state = any_state
    zone_val = None
    route_val = None

    new_truck_summary = TruckSummary(
        truckNumber=target_truck,
        customerName=customer_name,
        customerAddress=None,
        customerCity=customer_city,
        customerState=customer_state,
        zone=zone_val,
        route=route_val,
        totalWeight=float(total_weight),
        minWeight=int(min_weight),
        maxWeight=int(max_weight),
        totalOrders=int(total_orders),
        totalLines=int(total_lines),
        totalPieces=int(total_pieces),
        maxWidth=float(max_width),
        percentOverwidth=float(0.0),
        containsLate=bool(contains_late),
        priorityBucket=str(priority_bucket),
    )

    # Build updated assignments for changed trucks only
    changed_trucks = set(source_trucks_before) | {target_truck}
    updated_assignments_rows = assigns_df[assigns_df["truckNumber"].isin(
        list(changed_trucks))]
    updated_assignments: List[LineAssignment] = []
    for _, a in updated_assignments_rows.iterrows():
        tnum_val = a.get("truckNumber")
        try:
            tnum_int = int(tnum_val) if pd.notna(tnum_val) else target_truck
        except Exception:
            tnum_int = target_truck
        updated_assignments.append(LineAssignment(
            truckNumber=tnum_int,
            so=str(a.get("so")),
            line=str(a.get("line")),
            trttav_no=str(a.get("trttav_no")) if pd.notna(
                a.get("trttav_no")) else None,
            customerName=str(a.get("customerName")),
            customerAddress=a.get("customerAddress"),
            customerCity=str(a.get("customerCity")),
            customerState=str(a.get("customerState")),
            piecesOnTransport=int(a.get("piecesOnTransport") or 0),
            totalReadyPieces=int(a.get("totalReadyPieces") or 0),
            weightPerPiece=float(a.get("weightPerPiece") or 0.0),
            totalWeight=float(a.get("totalWeight") or 0.0),
            width=float(a.get("width") or 0.0),
            isOverwidth=bool(a.get("isOverwidth", False)),
            isLate=bool(a.get("isLate", False)),
            earliestDue=str(a.get("earliestDue")) if pd.notna(
                a.get("earliestDue")) else None,
            latestDue=str(a.get("latestDue")) if pd.notna(
                a.get("latestDue")) else None,
        ))

    return CombineTrucksResponse(
        success=True,
        message=f"Combined {len(selected_rows)} lines into truck {target_truck}",
        newTruck=new_truck_summary,
        updatedAssignments=updated_assignments,
        removedTruckIds=[int(i) for i in removed],
    )


@app.post("/export/dh-load-list")
async def export_dh_load_list(
    file: UploadFile = File(...),
    plannedDeliveryCol: Optional[str] = Form(None),
    planningWhse: Optional[str] = Form("ZAC"),
) -> StreamingResponse:
    """Export a DH Load List Excel per mapping.

    - Optional Planned Delivery column via 'plannedDeliveryCol' (UI now defaults to next business day)
    - Inserts a blue info row between transports (loads) with per-load stats in italics
    - Adds hidden blank column C to match provided layout
    """
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="Only .xlsx files are supported")

    try:
        content: bytes = await file.read()
        buffer = BytesIO(content)
        df: pd.DataFrame = pd.read_excel(buffer, engine="openpyxl")
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=400, detail=f"Failed to read Excel: {exc}") from exc

    # Apply Planning Whse filter first (default ZAC); if missing column, no-op
    if planningWhse:
        try:
            df = filter_by_planning_whse(df, allowed_values=(planningWhse,))
        except Exception:
            pass
    df = compute_calculated_fields(df)

    # Helper: compute next available business day (skip Sat/Sun)
    def next_business_day() -> Any:
        d = pd.Timestamp.today().normalize()
        d = d + pd.Timedelta(days=1)
        while d.weekday() >= 5:  # 5=Sat, 6=Sun
            d = d + pd.Timedelta(days=1)
        # return timezone-naive python datetime for Excel compatibility
        try:
            d = d.tz_localize(None)
        except Exception:
            pass
        return d.to_pydatetime()

    # Helper: normalize pandas/python datetimes to timezone-naive python datetime
    def as_dt(v: Any) -> Any:
        if isinstance(v, pd.Timestamp):
            try:
                if v.tz is not None or getattr(v, 'tzinfo', None) is not None:
                    v = v.tz_localize(None)
            except Exception:
                try:
                    v = v.tz_convert(None)
                except Exception:
                    pass
            return v.to_pydatetime()
        if isinstance(v, _dt.datetime):
            return v.replace(tzinfo=None) if v.tzinfo is not None else v
        return v

    # Helper: next business day after a given date (skip Sat/Sun)
    def next_business_day_after(value: Any) -> _dt.datetime:
        try:
            ts = pd.to_datetime(value, errors="coerce")
        except Exception:
            ts = pd.NaT
        if pd.isna(ts):
            return next_business_day()
        try:
            ts = ts.tz_localize(None)
        except Exception:
            try:
                ts = ts.tz_convert(None)
            except Exception:
                pass
        ts = ts + pd.Timedelta(days=1)
        while ts.weekday() >= 5:
            ts = ts + pd.Timedelta(days=1)
        return ts.to_pydatetime()

    # Helper: add N business days to a given date (skip Sat/Sun)
    def add_business_days(start: Any, days: int) -> _dt.datetime:
        try:
            ts = pd.to_datetime(start, errors="coerce")
        except Exception:
            ts = pd.NaT
        if pd.isna(ts):
            ts = pd.Timestamp.today().normalize()
        try:
            ts = ts.tz_localize(None)
        except Exception:
            try:
                ts = ts.tz_convert(None)
            except Exception:
                pass
        remaining = int(days or 0)
        step = 1 if remaining >= 0 else -1
        while remaining != 0:
            ts = ts + pd.Timedelta(days=step)
            if ts.weekday() < 5:
                remaining -= step
        return ts.to_pydatetime()

    # Planned Delivery column no longer drives Actual Ship; logic is per-load as specified
    has_planned_col = False

    # Optimize to get trucks and line splits (pieces/weights per transport)
    cfg = {"texas_max_lbs": 52000, "texas_min_lbs": 47000,
           "other_max_lbs": 48000, "other_min_lbs": 44000}
    try:
        trucks_df, assigns_df = naive_grouping(df.copy(), cfg)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=400, detail=f"Optimization failed: {exc}") from exc

    # Column normalization helpers
    def norm_key(s: Any) -> str:
        s = str(s)
        s = re.sub(r"\s+", " ", s)
        s = s.strip().lower()
        s = re.sub(r"[^a-z0-9]+", "", s)
        return s

    normalized = {norm_key(c): c for c in df.columns}

    def find_col(target: str, contains_ok: bool = True) -> Optional[str]:
        if target in normalized:
            return normalized[target]
        if contains_ok:
            for nk, orig in normalized.items():
                if target in nk:
                    return orig
        return None

    # Source columns
    col_type = find_col("type")
    col_bpcs = find_col("bpcs")  # BPcs totals per input line
    col_bal_weight = find_col("balweight") or find_col(
        "balanceweight") or find_col("bal weight")
    col_frm = find_col("frm")
    col_grd = find_col("grd")
    col_size = find_col("size")
    col_width = find_col("width") or "Width"
    col_lgth = find_col("lgth") or find_col("length")
    col_trttav = find_col(
        "trttavno", contains_ok=False) or find_col("trttavno")
    # Exact 'R' column from input for R# mapping
    col_r = find_col("r", contains_ok=False)
    col_latest_due = find_col("latestdue") or "Latest Due"
    col_earliest_due_src = find_col("earliestdue") or "Earliest Due"
    col_customer = find_col(
        "customer", contains_ok=False) or find_col("customer")
    col_so = find_col("so", contains_ok=False) or "SO"
    col_line = find_col("line", contains_ok=False) or "Line"
    whse_col = _find_planning_whse_col(df)

    # Index original rows by (SO, Line)
    keyed = df.copy()
    try:
        keyed["__so_key__"] = keyed[col_so].astype(str)
        keyed["__line_key__"] = keyed[col_line].astype(str)
    except Exception:
        keyed["__so_key__"] = keyed["SO"].astype(
            str) if "SO" in keyed.columns else pd.Series([None]*len(keyed), dtype="string")
        keyed["__line_key__"] = keyed["Line"].astype(
            str) if "Line" in keyed.columns else pd.Series([None]*len(keyed), dtype="string")
    index_map: Dict[tuple, Dict[str, Any]] = {}
    for _, r in keyed.iterrows():
        index_map[(str(r.get("__so_key__")), str(
            r.get("__line_key__")))] = r.to_dict()

    # Quick truck meta
    zones_by_truck: Dict[int, Optional[str]] = {}
    routes_by_truck: Dict[int, Optional[str]] = {}
    if not trucks_df.empty:
        for _, trow in trucks_df.iterrows():
            tnum = int(trow["truckNumber"]) if pd.notna(
                trow.get("truckNumber")) else None
            if tnum is None:
                continue
            zones_by_truck[tnum] = None if pd.isna(
                trow.get("zone")) else str(trow.get("zone"))
            routes_by_truck[tnum] = None if pd.isna(
                trow.get("route")) else str(trow.get("route"))

    headers = [
        "Actual Ship", "TR#", "Carrier", "Loaded", "Shipped", "Earliest Ship Date", "Ship Date", "Customer", "Type",
        "SO#", "SO Line", "R#", "WHSE", "Zone", "Route", "BPCS", "RPCS", "Bal Weight",
        "Ready Weight", "Frm", "Grd", "Size", "Width", "Lgth", "D", "PRV",
    ]

    # Build quick lookup for truck metadata
    trucks_meta: Dict[int, Dict[str, Any]] = {}
    if not trucks_df.empty:
        for _, t in trucks_df.iterrows():
            # Guard against missing/None truckNumber
            tnum_val = t.get("truckNumber")
            if tnum_val is None or (hasattr(pd, "isna") and pd.isna(tnum_val)):
                continue
            try:
                tnum = int(tnum_val)
            except Exception:  # noqa: BLE001
                continue
            trucks_meta[tnum] = {
                "totalWeight": float(t.get("totalWeight") or 0.0),
                "maxWeight": float(t.get("maxWeight") or 0.0),
                "containsLate": bool(t.get("containsLate", False)),
                "maxWidth": float(t.get("maxWidth") or 0.0),
            }
    # Map truck -> priority bucket based on its assignments
    bucket_by_truck: Dict[int, str] = {}
    if not assigns_df.empty:
        for tnum in assigns_df["truckNumber"].unique().tolist():
            try:
                tnum_i = int(tnum)
            except Exception:
                continue
            subset_tmp = assigns_df[assigns_df["truckNumber"] == tnum_i]
            bucket_val = "WithinWindow"
            if "priorityBucket" in subset_tmp.columns:
                vals = subset_tmp["priorityBucket"].dropna().astype(
                    str).tolist()
                if any(v == "Late" for v in vals):
                    bucket_val = "Late"
                elif any(v == "NearDue" for v in vals):
                    bucket_val = "NearDue"
            elif "isLate" in subset_tmp.columns and bool(subset_tmp["isLate"].any()):
                bucket_val = "Late"
            bucket_by_truck[tnum_i] = bucket_val
    # Helper to build rows for a given ordered list of trucks

    def build_rows_for(truck_order: List[int]) -> tuple[List[List[Any]], List[int], List[float]]:
        rows_local: List[List[Any]] = []
        sep_indices: List[int] = []
        sep_utils: List[float] = []

        def to_int(v: Any) -> Optional[int]:
            try:
                if v is None or (hasattr(pd, 'isna') and pd.isna(v)):
                    return None
                return int(float(v))
            except Exception:
                return None

        def to_float(v: Any) -> Optional[float]:
            try:
                if v is None or (hasattr(pd, 'isna') and pd.isna(v)):
                    return None
                return float(v)
            except Exception:
                return None
        for truck_num in truck_order:
            subset = assigns_df[assigns_df["truckNumber"] == truck_num]
            if subset.empty:
                continue
            # Compute per-load Actual Ship per rules:
            # - If any line is Late -> next business day
            # - Else -> day after latest Earliest Due among lines in this load
            contains_late_subset = bool(
                subset["isLate"].any()) if "isLate" in subset.columns else False
            if contains_late_subset:
                truck_actual_ship = as_dt(next_business_day())
            else:
                # Prefer assignments_df 'earliestDue'
                ed_series = None
                if "earliestDue" in subset.columns:
                    try:
                        ed_series = pd.to_datetime(
                            subset["earliestDue"], errors="coerce")
                    except Exception:
                        ed_series = None
                # Fallback to source 'Earliest Due' via index_map
                if ed_series is None or ed_series.dropna().empty:
                    ed_vals: list = []
                    for _, arow in subset.iterrows():
                        so = str(arow.get("so"))
                        line = str(arow.get("line"))
                        # Strip remainder suffixes for index_map lookup since source data doesn't have them
                        line_for_lookup = line
                        if line and line.endswith(('-R1', '-R2', '-R3', '-R4', '-R5', '-R6', '-R7', '-R8', '-R9')):
                            # Remove the -RX suffix
                            line_for_lookup = line[:-3]
                        src = index_map.get((so, line_for_lookup), {})
                        ed_vals.append(src.get(col_earliest_due_src))
                    try:
                        ed_series = pd.to_datetime(
                            pd.Series(ed_vals), errors="coerce")
                    except Exception:
                        ed_series = pd.Series([], dtype="datetime64[ns]")
                if ed_series.dropna().empty:
                    # No usable earliest due; default to next business day
                    truck_actual_ship = as_dt(next_business_day())
                else:
                    max_ed = ed_series.max()
                    # Next business day after max earliest due
                    truck_actual_ship = next_business_day_after(max_ed)
                    # If this date is in the past, move it 3 business days into the future from today
                    try:
                        tas = pd.to_datetime(
                            truck_actual_ship, errors="coerce")
                        today = pd.Timestamp.today().normalize()
                        if pd.notna(tas) and tas < today:
                            truck_actual_ship = add_business_days(today, 3)
                    except Exception:
                        pass
            for _, a in subset.iterrows():
                so = str(a.get("so"))
                line = str(a.get("line"))
                # Strip remainder suffixes (-R1, -R2, etc.) from line number for DH load list export
                line_for_export = line
                line_for_lookup = line
                if line and line.endswith(('-R1', '-R2', '-R3', '-R4', '-R5', '-R6', '-R7', '-R8', '-R9')):
                    # Remove the -RX suffix for export
                    line_for_export = line[:-3]
                    # Remove the -RX suffix for index_map lookup
                    line_for_lookup = line[:-3]
                cust = str(a.get("customerName"))
                src = index_map.get((so, line_for_lookup), {})
                ship_date = src.get(
                    col_latest_due) if col_latest_due in src else src.get("Latest Due")
                # as_dt defined above
                # Use on-transport values for counts and weights (no SO line quantities)
                bpcs_val = int(a.get("piecesOnTransport") or 0)
                bal_weight_val = float(a.get("totalWeight") or 0.0)
                # Frm: pull from source as-is (case-insensitive), do not force numeric
                frm_val = src.get(col_frm) if col_frm else None
                # keep as-is (text in data rows)
                grd_val = src.get(col_grd) if col_grd else None
                size_val = src.get(col_size) if col_size else None
                width_val = to_float(src.get(col_width)) if (
                    col_width and col_width in src) else to_float(a.get("width"))
                lgth_val = to_float(src.get(col_lgth)) if col_lgth else None
                d_val = src.get(col_trttav) if col_trttav else None
                # R#: from source 'R' column, if present
                rnum_val = None
                if col_r and col_r in src:
                    rnum_val = to_int(src.get(col_r))
                # PRV: 1 for data lines with weight
                prv_val = 1 if float(a.get("totalWeight")
                                     or 0.0) > 0.0 else None
                # Get earliest ship date from source data
                earliest_ship_date = src.get(
                    col_earliest_due_src) if col_earliest_due_src in src else src.get("Earliest Due")

                row = [
                    truck_actual_ship,
                    int(truck_num),
                    "Jordan Carriers",
                    "",
                    "",
                    as_dt(earliest_ship_date),
                    as_dt(ship_date),
                    cust,
                    src.get(col_type) if col_type else None,
                    so,
                    line_for_export,  # Use cleaned line number without remainder suffix
                    rnum_val,
                    src.get(whse_col) if whse_col else None,
                    zones_by_truck.get(int(truck_num)),
                    routes_by_truck.get(int(truck_num)),
                    bpcs_val,
                    int(a.get("piecesOnTransport") or 0),
                    bal_weight_val,
                    float(a.get("totalWeight") or 0.0),
                    frm_val,
                    grd_val,
                    size_val,
                    width_val,
                    lgth_val,
                    d_val,
                    prv_val,
                ]
                rows_local.append(row)
            # Info row
            total_ready_weight = float(subset["totalWeight"].sum()) if "totalWeight" in subset.columns else float(
                trucks_meta.get(int(truck_num), {}).get("totalWeight", 0.0))
            meta = trucks_meta.get(int(truck_num), {})
            max_weight = float(meta.get("maxWeight", 0.0))
            contains_late = bool(meta.get("containsLate", False)) or (
                bool(subset["isLate"].any()) if "isLate" in subset.columns else False)
            max_width = float(meta.get("maxWidth", 0.0))
            if "width" in subset.columns and not subset.empty:
                try:
                    max_width = max(float(x)
                                    for x in subset["width"].tolist() if pd.notna(x))
                except Exception:
                    pass
            pct_util = (total_ready_weight / max_weight *
                        100.0) if max_weight > 0 else 0.0
            late_status = "Late" if contains_late else "On time"
            overwidth_status = "Overwidth" if max_width > 96 else "Not Overwidth"
            sep_row = cast(List[Any], [None] * len(headers))
            for label, value in (("RPCS", late_status), ("Ready Weight", round(total_ready_weight, 2)), ("Frm", int(max_weight) if max_weight else None), ("Grd", f"{pct_util:.1f}%"), ("Width", overwidth_status)):
                try:
                    idx_col = headers.index(label)
                    sep_row[idx_col] = value
                except ValueError:
                    pass
            rows_local.append(sep_row)
            sep_indices.append(len(rows_local) - 1)
            sep_utils.append(float(pct_util))
        return rows_local, sep_indices, sep_utils

    # Determine sort order by percent utilization (descending)
    util_by_truck: Dict[int, float] = {}
    if not assigns_df.empty:
        for tnum in assigns_df["truckNumber"].unique().tolist():
            try:
                tnum_i = int(tnum)
            except Exception:
                continue
            subset_tmp = assigns_df[assigns_df["truckNumber"] == tnum_i]
            total_ready_weight_tmp = float(subset_tmp["totalWeight"].sum()) if "totalWeight" in subset_tmp.columns else float(
                trucks_meta.get(int(tnum_i), {}).get("totalWeight", 0.0))
            max_weight_tmp = float(trucks_meta.get(
                int(tnum_i), {}).get("maxWeight", 0.0))
            util_by_truck[tnum_i] = (
                total_ready_weight_tmp / max_weight_tmp) if max_weight_tmp > 0 else 0.0
    sorted_trucks = sorted(util_by_truck.keys(
    ), key=lambda k: util_by_truck.get(k, 0.0), reverse=True)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Define sheets and included buckets
        sheets_spec = [
            ("Late + NearDue", {"Late", "NearDue"}),
            ("WithinWindow", {"WithinWindow"}),
        ]
        for sheet, include in sheets_spec:
            truck_order = [tn for tn in sorted_trucks if bucket_by_truck.get(
                int(tn), "WithinWindow") in include]
            rows_local, sep_indices, sep_utils = build_rows_for(truck_order)
            out_df = pd.DataFrame(rows_local, columns=headers)
            out_df.to_excel(writer, sheet_name=sheet, index=False)

            ws = writer.sheets[sheet]
            # Insert hidden blank column C
            ws.insert_cols(3)
            ws.cell(row=1, column=3).value = None
            ws.column_dimensions[get_column_letter(3)].hidden = True
            # Freeze the header row
            ws.freeze_panes = "A2"

            # Bold header
            for c in range(1, ws.max_column + 1):
                ws.cell(row=1, column=c).font = Font(
                    name="Calibri", size=11, bold=True)

            # Shade and italicize only the separator/info rows in light blue
            fill = PatternFill(start_color="DCE6F1",
                               end_color="DCE6F1", fill_type="solid")
            for idx in sep_indices:
                excel_row = 2 + idx  # header is row 1
                if excel_row <= ws.max_row:
                    for c in range(1, ws.max_column + 1):
                        cell = ws.cell(row=excel_row, column=c)
                        cell.fill = fill
                        cell.font = Font(name="Calibri", size=11, italic=True)

            # Apply date format mm/dd/yyy to Actual Ship, Ship Date, and Earliest Ship Date columns
            header_to_col: Dict[str, int] = {}
            for c in range(1, ws.max_column + 1):
                header_val = ws.cell(row=1, column=c).value
                if isinstance(header_val, str):
                    header_to_col[header_val.strip()] = c
            date_fmt = "mm/dd/yyy"
            for header in ("Actual Ship", "Ship Date", "Earliest Ship Date"):
                cidx = header_to_col.get(header)
                if cidx:
                    for r in range(2, ws.max_row + 1):
                        cell = ws.cell(row=r, column=cidx)
                        cell.number_format = date_fmt

            # Color-code utilization % in separator rows (Grd column)
            grd_col = header_to_col.get("Grd")
            if grd_col and sep_indices:
                for idx, util in zip(sep_indices, sep_utils):
                    excel_row = 2 + idx
                    if excel_row <= ws.max_row:
                        cell = ws.cell(row=excel_row, column=grd_col)
                        # Keep italic; set color based on thresholds
                        if util >= 90.0:
                            color = "FF00B050"  # green
                        elif util >= 84.0:
                            color = "FFFFC000"  # yellow
                        else:
                            color = "FFFF0000"  # red
                        cell.font = Font(name="Calibri", size=11,
                                         italic=True, color=color)

            # Apply numeric formats to non-date numeric columns
            # Do not force numeric on Frm; it may be text in source
            int_cols = ["TR#", "R#", "BPCS", "RPCS", "PRV"]
            float_cols = ["Bal Weight", "Ready Weight", "Width", "Lgth"]
            # Treat RPCS specially: only format as number if the cell is numeric (for data rows, RPCS is not used; sep rows contain text)
            for hdr in int_cols:
                cidx = header_to_col.get(hdr)
                if not cidx:
                    continue
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=cidx)
                    if hdr == "RPCS" and isinstance(cell.value, str):
                        continue
                    cell.number_format = "0"
            for hdr in float_cols:
                cidx = header_to_col.get(hdr)
                if not cidx:
                    continue
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=cidx)
                    cell.number_format = "#,##0.00"

            # Standardize font to Calibri 11 for entire sheet while preserving bold/italic/colors
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    f = cell.font or Font()
                    cell.font = Font(name="Calibri", size=11,
                                     bold=f.bold, italic=f.italic, color=f.color)

            # Left-align all cells
            left_align = Alignment(horizontal="left")
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).alignment = left_align

            # Auto-fit column widths (skip hidden column C)
            for c in range(1, ws.max_column + 1):
                if ws.column_dimensions[get_column_letter(c)].hidden:
                    continue
                max_len = 0
                for r in range(1, ws.max_row + 1):
                    val = ws.cell(row=r, column=c).value
                    if val is None:
                        continue
                    sval = val if isinstance(val, str) else str(val)
                    if len(sval) > max_len:
                        max_len = len(sval)
                width = min(max_len + 2, 50)
                if width < 10:
                    width = 10
                ws.column_dimensions[get_column_letter(c)].width = width

    output.seek(0)
    return StreamingResponse(
        BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dh_load_list.xlsx"},
    )
